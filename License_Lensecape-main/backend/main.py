import datetime
import openpyxl
from flask import Flask, Response, jsonify
from flask_restful import Api, Resource
from capture import parse_arguments, start_capture
from object_detection import initialize_yolo_model, perform_object_detection, annotate_frame
from ocr import initialize_easyocr, perform_ocr, save_cropped_image
import supervision as sv
import cv2
from flask_cors import CORS

app = Flask(__name__)
api = Api(app)
CORS(app)

# Define constants
FRAME_RATE = 5  # Frames per second

def setup_workbook():
    try:
        LOG_FILE_PATH = 'vehicle_logs.xlsx'
        workbook = openpyxl.load_workbook(LOG_FILE_PATH)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Timestamp", "Plate Number", "Class ID", "Confidence"])

    return workbook.active

def read_logs(sheet):
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        timestamp, plate_number, class_id, confidence = row
        data.append({
            "Timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "Plate Number": plate_number,
            "Class ID": class_id,
            "Confidence": confidence
        })
    return data

def save_log(log_data, sheet):
    timestamp, plate_number, class_id, confidence = log_data
    plate_number = ''.join(char for char in plate_number if char.isprintable())
    sheet.append([timestamp, plate_number, class_id, confidence])

def generate_frames(cap, model, reader, save_folder, sheet):
    frame_interval = int(cap.get(cv2.CAP_PROP_FPS) / FRAME_RATE)
    frame_count = 0

    while True:
        ret, frame = cap.read()

        if frame_count % frame_interval == 0:
            result = perform_object_detection(model, frame)
            detections = sv.Detections.from_ultralytics(result)

            labels = []
            for confidence, class_id in zip(detections.confidence, detections.class_id):
                if class_id is not None:
                    label = f"{model.model.names[class_id]} {confidence:0.2f}"
                    if confidence > 0.7:
                        plate_image = frame[int(detections.xyxy[0][1]):int(detections.xyxy[0][3]),
                                           int(detections.xyxy[0][0]):int(detections.xyxy[0][2])]

                        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                        image_name = save_cropped_image(plate_image, save_folder, timestamp)

                        recognized_text = perform_ocr(reader, image_name)

                        if recognized_text:
                            label += f" Plate: {recognized_text}"

                            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            log_data = [timestamp, recognized_text, model.model.names[class_id], confidence]
                            save_log(log_data, sheet)
                    labels.append(label)
                else:
                    label = "Unknown"
                    labels.append(label)

            frame = annotate_frame(frame, detections, labels)
            ret, jpeg = cv2.imencode('.jpg', frame)

            if not ret:
                continue

            frame_bytes = jpeg.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n\r\n')

        frame_count += 1

@app.route('/api/video_feed_cam1')
def video_feed_cam1():
    print("Request received for /api/video_feed_cam1")
    return Response(generate_frames(cap, model, reader, save_folder, sheet),
                    mimetype='video/mp4',
                    content_type='multipart/x-mixed-replace; boundary=frame')

class VideoFeedCam1(Resource):
    def get(self):
        return Response(generate_frames(cap, model, reader, save_folder, sheet),
                        mimetype='image/jpeg',
                        content_type='multipart/x-mixed-replace; boundary=frame')

class ExcelDataResource(Resource):
    def get(self):
        sheet = setup_workbook()
        data = read_logs(sheet)
        return jsonify(data)

api.add_resource(VideoFeedCam1, '/api/video_feed_cam1')
api.add_resource(ExcelDataResource, '/api/get_excel_data')

if __name__ == "__main__":
    args = parse_arguments()

    try:
        cap = start_capture(args.webcam_resolution, args.camera_index)
    except cv2.error as e:
        print(f"Error: {e}")
        cap = None

    if cap is not None:
        model = initialize_yolo_model("/home/hakym/fyp_nust/model/best .pt")
        reader = initialize_easyocr()
        save_folder = 'cropped_images_cam1'
        sheet = setup_workbook()

        app.run(host='0.0.0.0', port=5000, threaded=True)
    else:
        print("Unable to start the application. Camera not found or accessible.")

